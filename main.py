#-*- coding=utf-8 -*-
import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


f = open("new.txt", 'w', errors='replace')
r = requests.get('https://apis.naver.com/cafe-home-web/cafe-home/v2/themecafes?page=1&perPage=18&sort=uppoint&type=ar&themeDir1Id=1&themeDir2Id=0')

data = json.loads(r.text)
themes = data['message']['result']['themes']
cafeList = []
sex_age_list = [
  {
    'sa': 'a10',
    'score': 0,
    'keyword': ['10대', '고등학생', '중학생']
  },
  {
    'sa': 'a20',
    'score': 0,
    'keyword': ['20대', '대학']
  },
  {
    'sa': 'a30',
    'score': 0,
    'keyword': ['30대']
  },
  {
    'sa': 'a40',
    'score': 0,
    'keyword': ['40대']
  },
  {
    'sa': 'a50',
    'score': 0,
    'keyword': ['50대']
  },
  {
    'sa': 'a60',
    'score': 0,
    'keyword': ['60대']
  },
  {
    'sa': 'a70',
    'score': 0,
    'keyword': ['70대']
  },
  {
    'sa': 'sm',
    'score': 0,
    'keyword': ['남성', '남자']
  },
  {
    'sa': 'sw',
    'score': 0,
    'keyword': ['여성', '여자']
  }
]

for j in themes:
  r = requests.get('https://apis.naver.com/cafe-home-web/cafe-home/v1/directories/'+str(j['themeId'])+'/sub-directories')
  subDirectories = json.loads(r.text)['message']['result']['directories']
  print(j['themeName']+' :'+str(len(subDirectories)))
  data = j['themeName']+' :'+str(len(subDirectories)) + '\n'
  f.write(data)

  for i in range(2):
    r = requests.get('https://apis.naver.com/cafe-home-web/cafe-home/v2/themecafes?page='+str(i+1)+'&perPage=50&sort=uppoint&type=ar&themeDir1Id='+str(j['themeId'])+'&themeDir2Id=0')
    cafeData = json.loads(r.text)['message']['result']['cafes']

    for item in cafeData:
      try:
        cafeList.append(
          {
            "cafeId": item['cafeId'],
            "cafeUrl": item['cafeUrl'],
            "cafeName": item['cafeName'],
            "introduction": item['introduction']
          }
        )
      except:
        cafeList.append(
          {
            "cafeId": item['cafeId'],
            "cafeUrl": item['cafeUrl'],
            "cafeName": item['cafeName'],
            "introduction": ''
          }
        )

    r = requests.get('https://apis.naver.com/cafe-home-web/cafe-home/v2/themecafes?page='+str(i+1)+'&perPage=50&sort=uppoint&type=at&themeDir1Id='+str(j['themeId'])+'&themeDir2Id=0')
    cafeData = json.loads(r.text)['message']['result']['cafes']

    for item in cafeData:
      try:
        cafeList.append(
          {
            "cafeId": item['cafeId'],
            "cafeUrl": item['cafeUrl'],
            "cafeName": item['cafeName'],
            "introduction": item['introduction']
          }
        )
      except:
        cafeList.append(
          {
            "cafeId": item['cafeId'],
            "cafeUrl": item['cafeUrl'],
            "cafeName": item['cafeName'],
            "introduction": ''
          }
        )
    
    for subDirItem in subDirectories:
      r = requests.get('https://apis.naver.com/cafe-home-web/cafe-home/v2/themecafes?page='+str(i+1)+'&perPage=50&sort=uppoint&type=ar&themeDir1Id='+str(j['themeId'])+'&themeDir2Id='+str(subDirItem['directoryId']))
      cafeData = json.loads(r.text)['message']['result']['cafes']
      
      for item in cafeData:
        try:
          cafeList.append(
            {
              "cafeId": item['cafeId'],
              "cafeUrl": item['cafeUrl'],
              "cafeName": item['cafeName'],
            "introduction": item['introduction']
            }
          )
        except:
          cafeList.append(
            {
              "cafeId": item['cafeId'],
              "cafeUrl": item['cafeUrl'],
              "cafeName": item['cafeName'],
              "introduction": ''
            }
          )
      
      r = requests.get('https://apis.naver.com/cafe-home-web/cafe-home/v2/themecafes?page='+str(i+1)+'&perPage=50&sort=uppoint&type=at&themeDir1Id='+str(j['themeId'])+'&themeDir2Id='+str(subDirItem['directoryId']))
      cafeData = json.loads(r.text)['message']['result']['cafes']
      
      for item in cafeData:
        try:
          cafeList.append(
            {
              "cafeId": item['cafeId'],
              "cafeUrl": item['cafeUrl'],
              "cafeName": item['cafeName'],
            "introduction": item['introduction']
            }
          )
        except:
          cafeList.append(
            {
              "cafeId": item['cafeId'],
              "cafeUrl": item['cafeUrl'],
              "cafeName": item['cafeName'],
              "introduction": ''
            }
          )
    


print(len(cafeList))
# json array 중복 제거
set_of_jsons = {json.dumps(d) for d in cafeList}
cafeList = [json.loads(d) for d in set_of_jsons]
print(len(cafeList))

for item in cafeList:
  for saItem in sex_age_list:
    for keyword in saItem['keyword']:
      if (item['cafeName'].find(keyword) != -1 or item['introduction'].find(keyword) != -1):
        saItem['score'] += 1
        break

  # if (item['cafeName'].find('10대') != -1 or item['introduction'].find('10대') != -1):
  #   sex_age_list['a10'] += 1
  # if (item['cafeName'].find('20대') != -1 or item['introduction'].find('20대') != -1):
  #   sex_age_list['a20'] += 1
  # if (item['cafeName'].find('30대') != -1 or item['introduction'].find('30대') != -1):
  #   sex_age_list['a30'] += 1
  # if (item['cafeName'].find('40대') != -1 or item['introduction'].find('40대') != -1):
  #   sex_age_list['a40'] += 1
  # if (item['cafeName'].find('50대') != -1 or item['introduction'].find('50대') != -1):
  #   sex_age_list['a50'] += 1
  # if (item['cafeName'].find('60대') != -1 or item['introduction'].find('60대') != -1):
  #   sex_age_list['a60'] += 1
  # if (item['cafeName'].find('70대') != -1 or item['introduction'].find('70대') != -1):
  #   sex_age_list['a70'] += 1
  # if (item['cafeName'].find('남성') != -1 or item['introduction'].find('남성') != -1 or item['cafeName'].find('남자') != -1 or item['introduction'].find('남자') != -1):
  #   sex_age_list['sm'] += 1
  # if (item['cafeName'].find('여성') != -1 or item['introduction'].find('여성') != -1 or item['cafeName'].find('여자') != -1 or item['introduction'].find('여자') != -1):
  #   sex_age_list['sw'] += 1
  
print(sex_age_list)

write_wb = Workbook()
write_ws = write_wb['Sheet']
write_ws.append(['#', 'cafeName', 'introduction'])
sheetName = 'Sheet'
count = 0
for item in cafeList:
  try:
    write_ws = write_wb[sheetName]
    write_ws.append([count+1, item['cafeName'], item['introduction']])
    fdata = str(count+1) + ' | ' + item['cafeName'] + ' : ' + item['introduction'] +'\n'
    f.write(fdata)
    count = count+1
    if (count % 5000 == 0):
      sheetName = 'Sheet' + str(count//5000)
      write_ws = write_wb.create_sheet(title=sheetName)
      write_ws = write_wb[sheetName]
      write_ws.append(['#', 'cafeName', 'introduction'])
  except :
    name = ILLEGAL_CHARACTERS_RE.sub(r'',item['cafeName'])
    intro = ILLEGAL_CHARACTERS_RE.sub(r'',item['introduction'])
    write_ws = write_wb[sheetName]
    write_ws.append([count+1, name, intro])
    fdata = str(count+1) + ' | ' + name + ' : ' + intro +'\n'
    f.write(fdata)
    count = count+1
    if (count % 5000 == 0):
      sheetName = 'Sheet' + str(count//5000)
      write_ws = write_wb.create_sheet(title=sheetName)
      write_ws = write_wb[sheetName]
      write_ws.append(['#','cafeName', 'introduction'])


write_wb.save('new.xlsx')
f.close()
print(count)