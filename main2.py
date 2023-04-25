from openpyxl import Workbook

write_wb = Workbook()
 
write_ws = write_wb.create_sheet('test')
 
#Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '숫자'

write_ws = write_wb['test']
#행 단위로 추가
write_ws.append([1,2,3])
 
#셀 단위로 추가
write_ws.cell(5,5,'5행5열')
write_wb.save('te11xt.xlsx')